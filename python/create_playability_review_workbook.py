from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parents[1]

INPUT_CSV = PROJECT_ROOT / "source" / "playability_review_queue.csv"
OUTPUT_XLSX = PROJECT_ROOT / "source" / "playability_review_queue.xlsx"


def read_csv_rows():
    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def main():
    rows = read_csv_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Playability Review"

    headers = [
        "GameID",
        "GameTitle",
        "RecommendationScore",
        "ReviewPriority",
        "ReviewSignal",
        "ReviewSignalCount",
        "ControlSchemeRisk",
        "SpecialAccessoryRequired",
        "RhythmFlag",
        "SHMUPFlag",
        "SoulslikeFlag",
        "MitigationAvailable",
        "PlayabilityNotes",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([
            int(row["GameID"]),
            row["GameTitle"],
            float(row["RecommendationScore"]),
            row["ReviewPriority"],
            row["ReviewSignal"],
            int(row["ReviewSignalCount"]),
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ])

    # ---------------------------------------------------------
    # Header formatting
    # ---------------------------------------------------------
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ---------------------------------------------------------
    # Column widths
    # ---------------------------------------------------------
    widths = {
        "A": 10,   # GameID
        "B": 42,   # GameTitle
        "C": 20,   # RecommendationScore
        "D": 16,   # ReviewPriority
        "E": 28,   # ReviewSignal
        "F": 18,   # ReviewSignalCount
        "G": 20,   # ControlSchemeRisk
        "H": 24,   # SpecialAccessoryRequired
        "I": 14,   # RhythmFlag
        "J": 14,   # SHMUPFlag
        "K": 14,   # SoulslikeFlag
        "L": 20,   # MitigationAvailable
        "M": 70,   # PlayabilityNotes
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ---------------------------------------------------------
    # Data validation
    # ---------------------------------------------------------
    control_risk_validation = DataValidation(
        type="list",
        formula1='"None,Moderate,High"',
        allow_blank=True,
    )

    boolean_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True,
    )

    ws.add_data_validation(control_risk_validation)
    ws.add_data_validation(boolean_validation)

    last_row = ws.max_row

    control_risk_validation.add(f"G2:G{last_row}")

    for col in ["H", "I", "J", "K", "L"]:
        boolean_validation.add(f"{col}2:{col}{last_row}")

    # ---------------------------------------------------------
    # Alignment
    # ---------------------------------------------------------
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # ---------------------------------------------------------
    # Instructions sheet
    # ---------------------------------------------------------
    instructions = wb.create_sheet("Instructions")

    instructions_rows = [
        ["GAIAS Playability Review"],
        [""],
        ["Purpose"],
        [
            "Use this workbook to manually review games prioritized by the "
            "GAIAS playability-enrichment queue."
        ],
        [""],
        ["ControlSchemeRisk"],
        ["None", "No identified material control-scheme risk."],
        [
            "Moderate",
            "Meaningful control, coordination, reaction, or legacy-control burden.",
        ],
        [
            "High",
            "Substantial control burden likely to materially affect playability.",
        ],
        [""],
        ["Boolean fields"],
        ["SpecialAccessoryRequired", "TRUE only when the game requires a special accessory."],
        ["RhythmFlag", "TRUE only when rhythm-focused gameplay is a defining requirement."],
        ["SHMUPFlag", "TRUE only when the game is genuinely a shoot-'em-up."],
        ["SoulslikeFlag", "TRUE when Souls-like difficulty/combat design is materially present."],
        [
            "MitigationAvailable",
            "TRUE when difficulty settings, assists, co-op, progression, alternate controls, "
            "or other meaningful mitigation exists.",
        ],
        [""],
        ["Important"],
        [
            "Review signals are prioritization aids only. Do not automatically classify "
            "a game based on Virtual Reality, Music, Shooter, or Side View taxonomy alone."
        ],
    ]

    for row in instructions_rows:
        instructions.append(row)

    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 90

    for cell in instructions[1]:
        cell.font = Font(bold=True, size=14)

    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUTPUT_XLSX)

    print(f"Created: {OUTPUT_XLSX}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()